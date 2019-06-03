from tkinter import *
from openpyxl import*
from tkinter.filedialog import askopenfilename
from PIL import Image, ImageDraw, ImageFont
 
window = Tk()
window.geometry('600x450')
window.title("Certificate Generator")
 
title = Label(window, text="CERTGEN", font=("Arial Bold",20))
scrollbar = Scrollbar(window, orient=VERTICAL)

choose_file = Label(window, text="Choose excel workbook")
choose_sheet = Label(window, text="choose sheet")
choose_column = Label(window, text="choose column")

#excel row

filename=""
sheetlist=[]

def chooseFile():
    global filename
    filename = askopenfilename()
    wb=load_workbook(filename)
    
    for name in wb.sheetnames:
        ws.insert(END, name)
        sheetlist.append(name)
        
wb_btn = Button(window, text="Choose Workbook",command=chooseFile) 

ws = Listbox(window, height=3,exportselection=0)
column_list = Listbox(window, height=3, yscrollcommand=scrollbar.set, exportselection=0)

sheet_name=""
active_index = 0
list_coords = {}
def onselect_sheet(evt):
    workbook=load_workbook(filename)
    w = evt.widget
    global sheet_name 
    sheet_name = w.get(w.curselection()[0])
    for i in range(0,len(sheetlist)):
        if(sheetlist[i]==sheet_name):
            workbook.active= i
            global active_index
            active_index = i
    for i in range(1,workbook.active.max_column+1):
        heading = workbook.active.cell(row=1, column = i)
        column_list.insert(END, heading.value)
        list_coords[heading.value]=[1, i]
        
ws.bind('<<ListboxSelect>>', onselect_sheet)

names = []

def col_select(evt):
    workbook=load_workbook(filename)
    workbook.active= active_index
    wsheet = workbook.active
    w = evt.widget
    col_name= w.get(w.curselection())
   #col_pos = list_coords[col_name][1]
    for i in range(2,wsheet.max_row+1):
        names.append(wsheet.cell(row=i, column=1).value)
       
        
column_list.bind('<<ListboxSelect>>', col_select)
scrollbar.config(command=column_list.yview)

label_family=Label(window, text="choose font file")
label_color=Label(window, text="font color in rgb")
label_size=Label(window, text="Enter font size")

font_file=""
def font_settings():
    global font_file
    font_file = askopenfilename()
    print(font_file)

def open_image():
    image_file=askopenfilename()
    image = Image.open(image_file)
    return image
    
choose_font_family= Button(window, text="choose font family", command=font_settings)
choose_font_color = Entry(window, text="enter color in RGB")
choose_font_color.insert(0,'rgb(0,0,0)')
choose_font_size = Entry(window, text="enter font size(pt)")

label_choose_image = Label(window,text="choose image file")
label_choose_format = Label(window, text="choose format")
label_choose_height = Label(window,text="enter height")

choose_image = Button(window, text="choose image",command=open_image)
choose_height= Entry(window, text="enter height")

OPTIONS = [
    "png",
    "pdf",
    "jpg"
]

formats = StringVar(window)
formats.set(OPTIONS[0]) # default value
choose_format=OptionMenu(window, formats, *OPTIONS)

def output():
    image = Image.open('cert.jpg')
    draw = ImageDraw.Draw(image)
    font_size= int(choose_font_size.get())
    font = ImageFont.truetype('Ananda Black Personal Use.ttf', size=font_size)
    #font = ImageFont.truetype('Ananda Black Personal Use.ttf', size=45)
    color = choose_font_color.get()
    height = int(choose_height.get())
    for name in range(0,len(names)):
        draw.text((50, height), names[name], fill=color, font=font)
        image.save('/users/ujwal/desktop/'+names[name]+'.jpg',"JPEG", quality=100, optimize=True, progressive=True)

Execute= Button(window,text="generate certificates",command=output)

#grid
title.grid(row=0,columnspan=3)
choose_file.grid(row=1,column=0, padx=40)
choose_sheet.grid(row=1,column=1, padx=40)
choose_column.grid(row=1,column=2, padx=40)

wb_btn.grid(row=2,column=0)
ws.grid(row=2,column=1)
column_list.grid(row=2, column=2,pady=20)

label_family.grid(row=3,column=0)
label_color.grid(row=3,column=1)
label_size.grid(row=3,column=2)

choose_font_family.grid(row=4,column=0)
choose_font_color.grid(row=4,column=1)
choose_font_size.grid(row=4,column=2, pady=20)

label_choose_image.grid(row=5,column=0)
label_choose_format.grid(row=5,column=1)
label_choose_height.grid(row=5,column=2)

choose_image.grid(row=6,column=0)
choose_format.grid(row=6,column=1)
choose_height.grid(row=6,column=2)

Execute.grid(row=7,column=0,rowspan=3,pady=50)

window.mainloop()
